from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Workbook configuration
# ---------------------------------------------------------------------------

TITLE = "Pollux — Cash Flow Statement"
SHEET_NAME = "Cash Flow Statement"

THIN_BORDER = Border(
    bottom=Side(style="thin")
)

TOTAL_BORDER = Border(
    top=Side(style="thin"),
    bottom=Side(style="double"),
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_int(value: Any) -> int | None:
    """
    Convert a year-like value into an integer.
    """
    if value is None:
        return None

    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _format_number(cell) -> None:
    """
    Apply financial number formatting.
    """
    cell.number_format = '#,##0;(#,##0);-'


def _style_total_row(ws, row_number: int) -> None:
    """
    Give total rows stronger visual separation.
    """
    for cell in ws[row_number]:
        cell.font = Font(bold=True)
        cell.border = TOTAL_BORDER


def _style_section_row(ws, row_number: int) -> None:
    """
    Style Operating / Investing / Financing section headers.
    """
    for cell in ws[row_number]:
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER


def _set_column_widths(ws) -> None:
    """
    Set practical widths for the financial statement.
    """
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------

def create_excel(
    statement: dict[str, Any],
    filename: str = "pollux_cash_flow_statement.xlsx",
) -> BytesIO:
    """
    Create an Excel workbook from Pollux's structured statement.

    The workbook is created entirely in memory. Nothing is written to disk.
    """

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME

    # ------------------------------------------------------------------
    # Metadata
    # ------------------------------------------------------------------

    current_year = _safe_int(statement.get("current_year"))
    previous_year = _safe_int(statement.get("previous_year"))

    worksheet.merge_cells("A1:D1")
    worksheet["A1"] = TITLE
    worksheet["A1"].font = Font(
        bold=True,
        size=16,
    )
    worksheet["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    worksheet.row_dimensions[1].height = 28

    worksheet.merge_cells("A2:D2")
    worksheet["A2"] = "Extracted by Pollux"
    worksheet["A2"].alignment = Alignment(
        horizontal="center",
    )
    worksheet["A2"].font = Font(
        italic=True,
    )

    # ------------------------------------------------------------------
    # Column headers
    # ------------------------------------------------------------------

    worksheet["A4"] = "Particulars"
    worksheet["B4"] = (
        str(current_year)
        if current_year is not None
        else "Current Year"
    )
    worksheet["C4"] = (
        str(previous_year)
        if previous_year is not None
        else "Previous Year"
    )
    worksheet["D4"] = "Status"

    for cell in worksheet[4]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        cell.border = THIN_BORDER

    # ------------------------------------------------------------------
    # Statement sections
    # ------------------------------------------------------------------

    sections = statement.get("sections", {})

    section_labels = {
        "operating": "Cash flows from operating activities",
        "investing": "Cash flows from investing activities",
        "financing": "Cash flows from financing activities",
    }

    current_row = 5

    for section_name in (
        "operating",
        "investing",
        "financing",
    ):
        section_rows = sections.get(section_name, [])

        # Section heading.
        worksheet.cell(
            row=current_row,
            column=1,
            value=section_labels[section_name],
        )

        worksheet.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=4,
        )

        _style_section_row(
            worksheet,
            current_row,
        )

        current_row += 1

        for item in section_rows:
            description = item.get("description", "")
            current_value = item.get("current")
            previous_value = item.get("previous")
            row_type = item.get("row_type", "data")

            worksheet.cell(
                row=current_row,
                column=1,
                value=description,
            )

            current_cell = worksheet.cell(
                row=current_row,
                column=2,
                value=current_value,
            )

            previous_cell = worksheet.cell(
                row=current_row,
                column=3,
                value=previous_value,
            )

            # Status is deliberately descriptive rather than corrective.
            status = "OK"

            if current_value is None or previous_value is None:
                status = "WARNING"

            worksheet.cell(
                row=current_row,
                column=4,
                value=status,
            )

            current_cell.alignment = Alignment(
                horizontal="right"
            )
            previous_cell.alignment = Alignment(
                horizontal="right"
            )

            _format_number(current_cell)
            _format_number(previous_cell)

            worksheet.cell(
                row=current_row,
                column=1,
            ).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )

            worksheet.cell(
                row=current_row,
                column=4,
            ).alignment = Alignment(
                horizontal="center",
            )

            if row_type == "total":
                _style_total_row(
                    worksheet,
                    current_row,
                )
            else:
                for cell in worksheet[current_row]:
                    cell.border = THIN_BORDER

            current_row += 1

        # Spacing between sections.
        current_row += 1

    # ------------------------------------------------------------------
    # Validation summary
    # ------------------------------------------------------------------

    validation = statement.get("validation", {})
    validation_status = validation.get(
        "status",
        "unknown",
    )
    warnings = validation.get(
        "warnings",
        [],
    )

    worksheet.cell(
        row=current_row,
        column=1,
        value="Validation status",
    )
    worksheet.cell(
        row=current_row,
        column=2,
        value=str(validation_status).upper(),
    )

    worksheet.cell(
        row=current_row,
        column=1,
    ).font = Font(bold=True)

    worksheet.cell(
        row=current_row,
        column=2,
    ).font = Font(bold=True)

    current_row += 1

    if warnings:
        worksheet.cell(
            row=current_row,
            column=1,
            value="Warnings",
        )
        worksheet.cell(
            row=current_row,
            column=1,
        ).font = Font(bold=True)

        current_row += 1

        for warning in warnings:
            worksheet.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=4,
            )

            worksheet.cell(
                row=current_row,
                column=1,
                value=f"• {warning}",
            )

            worksheet.cell(
                row=current_row,
                column=1,
            ).alignment = Alignment(
                wrap_text=True,
            )

            current_row += 1

    # ------------------------------------------------------------------
    # Worksheet usability
    # ------------------------------------------------------------------

    worksheet.freeze_panes = "A5"

    worksheet.sheet_view.showGridLines = False

    worksheet.auto_filter.ref = (
        f"A4:D{max(4, current_row - 1)}"
    )

    _set_column_widths(worksheet)

    # Vertical alignment for used cells.
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=current_row,
        min_col=1,
        max_col=4,
    ):
        for cell in row:
            cell.alignment = Alignment(
                horizontal=(
                    cell.alignment.horizontal
                    or "left"
                ),
                vertical="center",
                wrap_text=cell.alignment.wrap_text,
            )

    # ------------------------------------------------------------------
    # Return in-memory workbook
    # ------------------------------------------------------------------

    output = BytesIO()

    workbook.save(output)
    output.seek(0)

    return output