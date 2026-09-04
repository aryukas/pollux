from io import BytesIO

from openpyxl import load_workbook

from excel import create_excel


def build_test_statement():
    return {
        "statement_type": "cash_flow_statement",
        "current_year": 2023,
        "previous_year": 2022,
        "sections": {
            "operating": [
                {
                    "description": "Profit before taxation",
                    "current": 10473,
                    "previous": 10194,
                    "section": "operating",
                    "row_type": "data",
                },
                {
                    "description": "Depreciation and amortization expense",
                    "current": 1533,
                    "previous": 1194,
                    "section": "operating",
                    "row_type": "data",
                },
                {
                    "description": "Cash Generated from/(Used in) Operations",
                    "current": 16780,
                    "previous": 2129,
                    "section": "operating",
                    "row_type": "total",
                },
                {
                    "description": "Direct Taxes Paid",
                    "current": -2820,
                    "previous": -2820,
                    "section": "operating",
                    "row_type": "data",
                },
                {
                    "description": "Net cash from operating activities",
                    "current": 25966,
                    "previous": 10697,
                    "section": "operating",
                    "row_type": "total",
                },
            ],
            "investing": [
                {
                    "description": "Purchase of property, plant and equipment",
                    "current": -2084,
                    "previous": -3673,
                    "section": "investing",
                    "row_type": "data",
                },
                {
                    "description": "Net cash from investing activities",
                    "current": -2084,
                    "previous": -3073,
                    "section": "investing",
                    "row_type": "total",
                },
            ],
            "financing": [
                {
                    "description": "Proceeds from long term borrowings",
                    "current": -400,
                    "previous": 1669,
                    "section": "financing",
                    "row_type": "data",
                },
                {
                    "description": "Proceeds from short term borrowings",
                    "current": 4,
                    "previous": 436,
                    "section": "financing",
                    "row_type": "data",
                },
                {
                    "description": "Net cash from financing activities",
                    "current": -396,
                    "previous": 2104,
                    "section": "financing",
                    "row_type": "total",
                },
            ],
        },
        "validation": {
            "status": "valid",
            "warnings": [],
        },
    }


def test_excel_export():
    statement = build_test_statement()

    output = create_excel(statement)

    # ---------------------------------------------------------------
    # Basic output validation
    # ---------------------------------------------------------------

    assert isinstance(output, BytesIO)

    data = output.getvalue()

    assert len(data) > 0

    # XLSX files are ZIP-based.
    assert data[:2] == b"PK"

    # ---------------------------------------------------------------
    # Load generated workbook
    # ---------------------------------------------------------------

    output.seek(0)

    workbook = load_workbook(output)

    assert "Cash Flow Statement" in workbook.sheetnames

    worksheet = workbook["Cash Flow Statement"]

    # ---------------------------------------------------------------
    # Title / metadata
    # ---------------------------------------------------------------

    assert worksheet["A1"].value == "Pollux — Cash Flow Statement"
    assert worksheet["A2"].value == "Extracted by Pollux"

    # ---------------------------------------------------------------
    # Headers
    # ---------------------------------------------------------------

    assert worksheet["A4"].value == "Particulars"
    assert worksheet["B4"].value == "2023"
    assert worksheet["C4"].value == "2022"
    assert worksheet["D4"].value == "Status"

    # ---------------------------------------------------------------
    # Locate rows by description
    # ---------------------------------------------------------------

    rows_by_description = {}

    for row in worksheet.iter_rows(
        min_row=5,
        max_col=4,
    ):
        description = row[0].value

        if description:
            rows_by_description[description] = row

    # ---------------------------------------------------------------
    # Verify operating data
    # ---------------------------------------------------------------

    profit = rows_by_description["Profit before taxation"]

    assert profit[1].value == 10473
    assert profit[2].value == 10194
    assert profit[3].value == "OK"

    taxes = rows_by_description["Direct Taxes Paid"]

    assert taxes[1].value == -2820
    assert taxes[2].value == -2820

    operating_total = rows_by_description[
        "Net cash from operating activities"
    ]

    assert operating_total[1].value == 25966
    assert operating_total[2].value == 10697

    # ---------------------------------------------------------------
    # Verify investing
    # ---------------------------------------------------------------

    ppe = rows_by_description[
        "Purchase of property, plant and equipment"
    ]

    assert ppe[1].value == -2084
    assert ppe[2].value == -3673

    investing_total = rows_by_description[
        "Net cash from investing activities"
    ]

    assert investing_total[1].value == -2084
    assert investing_total[2].value == -3073

    # ---------------------------------------------------------------
    # Verify financing
    # ---------------------------------------------------------------

    long_term = rows_by_description[
        "Proceeds from long term borrowings"
    ]

    assert long_term[1].value == -400
    assert long_term[2].value == 1669

    financing_total = rows_by_description[
        "Net cash from financing activities"
    ]

    assert financing_total[1].value == -396
    assert financing_total[2].value == 2104

    # ---------------------------------------------------------------
    # Verify negative number formatting
    # ---------------------------------------------------------------

    assert taxes[1].number_format == "#,##0;(#,##0);-"
    assert ppe[1].number_format == "#,##0;(#,##0);-"

    # ---------------------------------------------------------------
    # Verify total-row formatting
    # ---------------------------------------------------------------

    assert operating_total[0].font.bold is True
    assert investing_total[0].font.bold is True
    assert financing_total[0].font.bold is True

    # ---------------------------------------------------------------
    # Verify worksheet usability
    # ---------------------------------------------------------------

    assert worksheet.freeze_panes == "A5"
    assert worksheet.sheet_view.showGridLines is False

    # ---------------------------------------------------------------
    # Verify all three sections exist
    # ---------------------------------------------------------------

    assert "Cash flows from operating activities" in rows_by_description
    assert "Cash flows from investing activities" in rows_by_description
    assert "Cash flows from financing activities" in rows_by_description

    print("\n" + "=" * 90)
    print("ALL EXCEL TESTS PASSED")
    print("=" * 90)


if __name__ == "__main__":
    test_excel_export()